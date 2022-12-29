import csv
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00


def do_exit(message: str):
    print(message)
    exit()


class UserInput:
    def __init__(self):
        self.file_name = input("Введите название файла: ")
        self.profession = input("Введите название профессии: ")


class Salary:
    salary_from: str
    salary_to: str
    salary_currency: str
    average_salary: int
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def set_average_salary(self):
        self.average_salary = int(self.currency_to_rub[self.salary_currency] *
                                  (float(self.salary_from) + float(self.salary_to)) // 2)

    def set_attribute(self, key: str, value: str):
        if key == "salary_gross":
            return
        if key != 'salary_currency':
            value = float(value)
        self.__setattr__(key, value)

class DataSet:
    def csv_reader(file_name: str):
        vacancies_array = []
        is_empty_file = True
        with open(file_name, 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            if not reader.fieldnames:
                do_exit("Пустой файл")
            for row in reader:
                is_empty_file = False
                if all(row.values()):
                    vacancy = Vacancy(row)
                    vacancies_array.append(vacancy)
        if is_empty_file:
            do_exit("Нет данных")
        return vacancies_array


class Vacancy:
    name: str
    employer_name: str
    salary: Salary
    area_name: str
    published_at: int

    def __init__(self, fields: dict):
        self.date_time_publishing = None
        for key, value in fields.items():
            value = self.delete_html_tags(value)
            if not self.check_salary(key, value):
                self.__setattr__(key, value)

        self.published_time_formatter()
        self.salary.set_average_salary()

    def delete_html_tags(self, value: str) -> str:
        value = re.sub('<.*?>', '', str(value)).replace("\r\n", "\n")
        value = ' '.join(value.split()).strip()
        return value

    def check_salary(self, key: str, value: str) -> bool:
        if not key.__contains__("salary"):
            return False
        if not hasattr(self, "salary"):
            self.salary = Salary()
        self.salary.set_attribute(key, value)
        return True

    def published_time_formatter(self):
        hour, minute, second = self.published_at.split('T')[1].split('+')[0].split(':')
        year, month, day = self.published_at.split('T')[0].split('-')
        self.date_time_publishing = datetime(int(year), int(month), int(day), int(hour), int(minute), int(second))
        self.published_at = int(year)

    def get_field(self, field: str):
        if field == 'salary':
            return self.salary.average_salary
        return self.__getattribute__(field)


class DataDictionary:
    def __init__(self):
        self.salary_years = {} #a
        self.vacancies_years = {} #b
        self.salary_years_by_profession = {} #c
        self.vacancies_years_by_profession = {} #d
        self.salaries_cities = {} #e
        self.vacancy_cities_ratio = {} #f
        self.city_vacancies_count = {}



    def update_data_by_vacancy(self, vacancy, profession: str):
        self.update_vacancies_count_dict('city_vacancies_count', 'area_name', vacancy)
        self.update_salary_dict('salary_years', 'published_at', vacancy)
        self.update_vacancies_count_dict('vacancies_years', 'published_at', vacancy)
        self.update_salary_dict('salaries_cities', 'area_name', vacancy)
        self.update_vacancies_count_dict('vacancy_cities_ratio', 'area_name', vacancy)
        if vacancy.name.__contains__(profession):
            self.update_salary_dict('salary_years_by_profession', 'published_at', vacancy)
            self.update_vacancies_count_dict('vacancies_years_by_profession', 'published_at', vacancy)

    def update_salary_dict(self, dict_name: str, field: str, vac: Vacancy) -> None:
        dictionary = self.__getattribute__(dict_name)
        key = vac.get_field(field)
        if key not in dictionary.keys():
            dictionary[key] = [vac.salary.average_salary, 1]
        else:
            dictionary[key][0] += vac.salary.average_salary
            dictionary[key][1] += 1

    def update_data(self, vacancies: list, profession: str) -> None:
        self.profession = profession
        for vacancy in vacancies:
            self.update_data_by_vacancy(vacancy, profession)

        self.correct_data(vacancies)

    def get_first(self, dictionary: dict, vacancies: list, amount: int) -> dict:
        count = 0
        res = {}
        for key, value in dictionary.items():
            if count == amount:
                break
            if self.city_vacancies_count[key] >= len(vacancies) // 100:
                res[key] = value
                count += 1
        return res

    def update_vacancies_count_dict(self, dict_name: str, field: str, vac: Vacancy) -> None:
        dictionary = self.__getattribute__(dict_name)
        key = vac.get_field(field)
        if key not in dictionary.keys():
            dictionary[key] = 1
        else:
            dictionary[key] += 1

    def correct_data(self, vacancies: list):
        for key, value in self.vacancy_cities_ratio.items():
            self.vacancy_cities_ratio[key] = round(value / len(vacancies), 4)

        buf = dict(sorted(self.salaries_cities.items(), key=lambda x: x[1][1] / x[1][0]))
        self.salaries_cities = self.get_first(buf, vacancies, 10)

        buf = dict(sorted(self.vacancy_cities_ratio.items(), key=lambda x: x[1], reverse=True))
        self.vacancy_cities_ratio = self.get_first(buf, vacancies, 10)



    def print(self) -> None:
        print_dictionary: {str, dict} = {
            "Динамика уровня зарплат по годам: ": self.salary_years,
            "Динамика количества вакансий по годам: ": self.vacancies_years,
            "Динамика уровня зарплат по годам для выбранной профессии: ": self.salary_years_by_profession,
            "Динамика количества вакансий по годам для выбранной профессии: ": self.vacancies_years_by_profession,
            "Уровень зарплат по городам (в порядке убывания): ": self.salaries_cities,
            "Доля вакансий по городам (в порядке убывания): ": self.vacancy_cities_ratio
        }
        for key, value in print_dictionary.items():
            if len(value) == 0:
                value = {k: 0 for k in self.salary_years.keys()}
            for k, v in value.items():
                if type(v) is list:
                    value[k] = v[0] // v[1]
            print(f"{key}{value}")





class Report:
    wb: Workbook

    def __init__(self, data: DataDictionary):
        self.data = data
        self.work_book = Workbook()

    def generate_excel(self):
        self.work_book.remove(self.work_book.active)
        self.generate_statistics_by_years()
        self.generate_statistics_by_cities()
        self.work_book.save("report.xlsx")

    def generate_statistics_by_years(self):
        ws = self.work_book.create_sheet("Статистика по годам")
        self.generate_data_dictionary(ws, "A", "Год", {v: str(k) for k, v in data.salary_years.items()})
        self.generate_data_dictionary(ws, "B", "Средняя зарплата", data.salary_years)
        self.generate_data_dictionary(ws, "C",
                                   f"Средняя зарплата - {data.profession}", data.salary_years_by_profession)
        self.generate_data_dictionary(ws, "D", "Количество вакансий", data.vacancies_years)
        self.generate_data_dictionary(ws, "E",
                                   f"Количество вакансий - {data.profession}", data.vacancies_years_by_profession)
        self.update_cell_settings(ws)

    def generate_data_dictionary(self, ws, column: str, name: str, dictionary: dict):
        ws[f"{column}1"] = name
        count = 2
        for year, value in dictionary.items():
            ws[f"{column}{count}"] = value
            count += 1

    def generate_statistics_by_cities(self):
        ws = self.work_book.create_sheet("Статистика по городам")
        self.generate_data_dictionary(ws, "A",
                                            "Город", {v: k for k, v in data.salaries_cities.items()})
        self.generate_data_dictionary(ws, "B", "Уровень зарплат", data.salaries_cities)
        self.generate_data_dictionary(ws, "D",
                                            "Город", {v: k for k, v in data.vacancy_cities_ratio.items()})
        self.generate_data_dictionary(ws, "E",
                                      "Доля вакансий", data.vacancy_cities_ratio)
        self.set_percent_style(ws)
        self.update_cell_settings(ws)

    def update_cell_settings(self, ws):
        self.set_cell(ws)
        self.set_correctly_column_width(ws)

    def set_percent_style(self, ws):
        for i in range(2, 12):
            ws[f"E{i}"].number_format = FORMAT_PERCENTAGE_00



    def set_cell(self, ws):
        isFirstCell = True
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    cell.border = Border(top=Side(border_style="thin", color="000000"),
                                         left=Side(border_style="thin", color="000000"),
                                         right=Side(border_style="thin", color="000000"),
                                         bottom=Side(border_style="thin", color="000000"))
                    if isFirstCell:
                        cell.font = Font(bold=True)
            isFirstCell = False

    def set_correctly_column_width(self, ws):
        a = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E", 6: "F", 7: "G"}
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)) + 1))

        for col, value in dims.items():
            ws.column_dimensions[a[col - 1]].width = value


user_input = UserInput()
vacancies_array = DataSet.csv_reader(user_input.file_name)

if len(vacancies_array) == 0:
    do_exit("Ничего не найдено")

data = DataDictionary()
data.update_data(vacancies_array, user_input.profession)
data.print()

report = Report(data)
report.generate_excel()